#!/usr/bin/env python3
"""
Crea el proyecto Tableau: Mundiales por Confederaciones
Genera:
  - Mundiales_Datos.xlsx   (data source limpio)
  - Mundiales_Confederaciones.twbx  (packaged Tableau workbook)
"""

import sys, os, zipfile, io
sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = r"F:\LUEGO BORRAR\PROYECTO FINAL"
EXCEL_NAME = "Mundiales_Datos.xlsx"
TWBX_NAME  = "Mundiales_Confederaciones.twbx"
TWB_NAME   = "Mundiales_Confederaciones.twb"

# ─────────────────────────────────────────────────────────────────────────────
# 1.  DATOS BRUTOS
# ─────────────────────────────────────────────────────────────────────────────

# Cada edición: (año, sede, n_equipos, campeón, subcampeón, 3ro, 4to)
EDICIONES_RAW = [
    (1930,"Uruguay",     13,"Uruguay",       "Argentina",      "Estados Unidos","Yugoslavia"),
    (1934,"Italia",      16,"Italia",         "Checoeslovaquia","Alemania",      "Austria"),
    (1938,"Francia",     15,"Italia",         "Hungría",        "Brasil",        "Suecia"),
    (1950,"Brasil",      13,"Uruguay",        "Brasil",         "Suecia",        "España"),
    (1954,"Suiza",       16,"Alemania",       "Hungría",        "Austria",       "Uruguay"),
    (1958,"Suecia",      16,"Brasil",         "Suecia",         "Francia",       "Alemania"),
    (1962,"Chile",       16,"Brasil",         "Checoeslovaquia","Chile",         "Yugoslavia"),
    (1966,"Inglaterra",  16,"Inglaterra",     "Alemania",       "Portugal",      "Unión Soviética"),
    (1970,"México",      16,"Brasil",         "Italia",         "Alemania",      "Uruguay"),
    (1974,"Alemania",    16,"Alemania",       "Países Bajos",   "Polonia",       "Brasil"),
    (1978,"Argentina",   16,"Argentina",      "Países Bajos",   "Brasil",        "Italia"),
    (1982,"España",      24,"Italia",         "Alemania",       "Polonia",       "Francia"),
    (1986,"México",      24,"Argentina",      "Alemania",       "Francia",       "Bélgica"),
    (1990,"Italia",      24,"Alemania",       "Argentina",      "Italia",        "Inglaterra"),
    (1994,"Estados Unidos",24,"Brasil",       "Italia",         "Suecia",        "Bulgaria"),
    (1998,"Francia",     32,"Francia",        "Brasil",         "Croacia",       "Países Bajos"),
    (2002,"Corea/Japón", 32,"Brasil",         "Alemania",       "Turquía",       "Corea del Sur"),
    (2006,"Alemania",    32,"Italia",         "Francia",        "Alemania",      "Portugal"),
    (2010,"Sudáfrica",   32,"España",         "Países Bajos",   "Alemania",      "Uruguay"),
    (2014,"Brasil",      32,"Alemania",       "Argentina",      "Países Bajos",  "Brasil"),
    (2018,"Rusia",       32,"Francia",        "Croacia",        "Bélgica",       "Inglaterra"),
    (2022,"Catar",       32,"Argentina",      "Francia",        "Croacia",       "Marruecos"),
]

# Confederación de cada país participante histórico
CONF = {
    "Alemania":          ("UEFA","Europa/UEFA","DEU"),
    "Alemania Democrática":("UEFA","Europa/UEFA","DDR"),
    "Angola":            ("CAF","África/CAF","AGO"),
    "Arabia Saudita":    ("AFC","Asia/AFC","SAU"),
    "Argelia":           ("CAF","África/CAF","DZA"),
    "Argentina":         ("CONMEBOL","América Sur/CONMEBOL","ARG"),
    "Australia":         ("AFC","Asia/AFC","AUS"),
    "Austria":           ("UEFA","Europa/UEFA","AUT"),
    "Bélgica":           ("UEFA","Europa/UEFA","BEL"),
    "Bolivia":           ("CONMEBOL","América Sur/CONMEBOL","BOL"),
    "Bosnia y Herzegovina":("UEFA","Europa/UEFA","BIH"),
    "Brasil":            ("CONMEBOL","América Sur/CONMEBOL","BRA"),
    "Bulgaria":          ("UEFA","Europa/UEFA","BGR"),
    "Camerún":           ("CAF","África/CAF","CMR"),
    "Canadá":            ("CONCACAF","América Norte/CONCACAF","CAN"),
    "Catar":             ("AFC","Asia/AFC","QAT"),
    "Chile":             ("CONMEBOL","América Sur/CONMEBOL","CHL"),
    "China":             ("AFC","Asia/AFC","CHN"),
    "Colombia":          ("CONMEBOL","América Sur/CONMEBOL","COL"),
    "Corea del Norte":   ("AFC","Asia/AFC","PRK"),
    "Corea del Sur":     ("AFC","Asia/AFC","KOR"),
    "Costa de Marfil":   ("CAF","África/CAF","CIV"),
    "Costa Rica":        ("CONCACAF","América Norte/CONCACAF","CRI"),
    "Croacia":           ("UEFA","Europa/UEFA","HRV"),
    "Cuba":              ("CONCACAF","América Norte/CONCACAF","CUB"),
    "Checoeslovaquia":   ("UEFA","Europa/UEFA","TCH"),
    "Dinamarca":         ("UEFA","Europa/UEFA","DNK"),
    "Ecuador":           ("CONMEBOL","América Sur/CONMEBOL","ECU"),
    "Egipto":            ("CAF","África/CAF","EGY"),
    "El Salvador":       ("CONCACAF","América Norte/CONCACAF","SLV"),
    "Emiratos Árabes Unidos":("AFC","Asia/AFC","ARE"),
    "Escocia":           ("UEFA","Europa/UEFA","SCO"),
    "Eslovaquia":        ("UEFA","Europa/UEFA","SVK"),
    "Eslovenia":         ("UEFA","Europa/UEFA","SVN"),
    "España":            ("UEFA","Europa/UEFA","ESP"),
    "Estados Unidos":    ("CONCACAF","América Norte/CONCACAF","USA"),
    "Francia":           ("UEFA","Europa/UEFA","FRA"),
    "Gales":             ("UEFA","Europa/UEFA","WAL"),
    "Ghana":             ("CAF","África/CAF","GHA"),
    "Grecia":            ("UEFA","Europa/UEFA","GRC"),
    "Haití":             ("CONCACAF","América Norte/CONCACAF","HTI"),
    "Honduras":          ("CONCACAF","América Norte/CONCACAF","HND"),
    "Hungría":           ("UEFA","Europa/UEFA","HUN"),
    "Indonesia":         ("AFC","Asia/AFC","IDN"),
    "Inglaterra":        ("UEFA","Europa/UEFA","ENG"),
    "Irán":              ("AFC","Asia/AFC","IRN"),
    "Irak":              ("AFC","Asia/AFC","IRQ"),
    "Irlanda":           ("UEFA","Europa/UEFA","IRL"),
    "Irlanda del Norte": ("UEFA","Europa/UEFA","NIR"),
    "Islandia":          ("UEFA","Europa/UEFA","ISL"),
    "Israel":            ("UEFA","Europa/UEFA","ISR"),
    "Italia":            ("UEFA","Europa/UEFA","ITA"),
    "Jamaica":           ("CONCACAF","América Norte/CONCACAF","JAM"),
    "Japón":             ("AFC","Asia/AFC","JPN"),
    "Kuwait":            ("AFC","Asia/AFC","KWT"),
    "Marruecos":         ("CAF","África/CAF","MAR"),
    "México":            ("CONCACAF","América Norte/CONCACAF","MEX"),
    "Nigeria":           ("CAF","África/CAF","NGA"),
    "Noruega":           ("UEFA","Europa/UEFA","NOR"),
    "Nueva Zelanda":     ("OFC","Oceanía/OFC","NZL"),
    "Países Bajos":      ("UEFA","Europa/UEFA","NLD"),
    "Panamá":            ("CONCACAF","América Norte/CONCACAF","PAN"),
    "Paraguay":          ("CONMEBOL","América Sur/CONMEBOL","PRY"),
    "Perú":              ("CONMEBOL","América Sur/CONMEBOL","PER"),
    "Polonia":           ("UEFA","Europa/UEFA","POL"),
    "Portugal":          ("UEFA","Europa/UEFA","PRT"),
    "República Checa":   ("UEFA","Europa/UEFA","CZE"),
    "República Democrática del Congo":("CAF","África/CAF","COD"),
    "Rumania":           ("UEFA","Europa/UEFA","ROU"),
    "Rusia":             ("UEFA","Europa/UEFA","RUS"),
    "Senegal":           ("CAF","África/CAF","SEN"),
    "Serbia":            ("UEFA","Europa/UEFA","SRB"),
    "Sudáfrica":         ("CAF","África/CAF","ZAF"),
    "Suecia":            ("UEFA","Europa/UEFA","SWE"),
    "Suiza":             ("UEFA","Europa/UEFA","CHE"),
    "Togo":              ("CAF","África/CAF","TGO"),
    "Trinidad y Tobago": ("CONCACAF","América Norte/CONCACAF","TTO"),
    "Túnez":             ("CAF","África/CAF","TUN"),
    "Turquía":           ("UEFA","Europa/UEFA","TUR"),
    "Ucrania":           ("UEFA","Europa/UEFA","UKR"),
    "Unión Soviética":   ("UEFA","Europa/UEFA","URS"),
    "Uruguay":           ("CONMEBOL","América Sur/CONMEBOL","URY"),
    "Yugoslavia":        ("UEFA","Europa/UEFA","YUG"),
    "Zaire":             ("CAF","África/CAF","ZAR"),
    # Variantes ortográficas históricas
    "Checoslovaquia":       ("UEFA","Europa/UEFA","TCH"),
    "Checoeslovaquia":      ("UEFA","Europa/UEFA","TCH"),
    "República de Irlanda": ("UEFA","Europa/UEFA","IRL"),
    "Iraq":                 ("AFC","Asia/AFC","IRQ"),
    "Serbia y Montenegro":  ("UEFA","Europa/UEFA","SCG"),
    "Emiratos Árabes":      ("AFC","Asia/AFC","ARE"),
}

# Partidos por edición: {año: [(pais, fase), ...]}
# Fase: 'Campeón'=6, 'Subcampeón'=5, 'Tercero'=4, 'Cuarto'=3,
#       'Cuartos de Final'=2, 'Octavos de Final'=1, 'Fase de Grupos'=0
FASE_NUM = {
    'Campeón':6, 'Subcampeón':5, 'Tercero':4, 'Cuarto':3,
    'Cuartos de Final':2, 'Octavos de Final':1, 'Fase de Grupos':0
}

PARTICIPACIONES_RAW = {
1930:[
    ("Uruguay","Campeón"),("Argentina","Subcampeón"),("Estados Unidos","Tercero"),
    ("Yugoslavia","Cuarto"),("Brasil","Fase de Grupos"),("Francia","Fase de Grupos"),
    ("Chile","Fase de Grupos"),("México","Fase de Grupos"),("Bolivia","Fase de Grupos"),
    ("Paraguay","Fase de Grupos"),("Perú","Fase de Grupos"),("Bélgica","Fase de Grupos"),
    ("Rumania","Fase de Grupos"),
],
1934:[
    ("Italia","Campeón"),("Checoeslovaquia","Subcampeón"),("Alemania","Tercero"),
    ("Austria","Cuarto"),("España","Cuartos de Final"),("Hungría","Cuartos de Final"),
    ("Suiza","Cuartos de Final"),("Suecia","Cuartos de Final"),
    ("Estados Unidos","Octavos de Final"),("Brasil","Octavos de Final"),
    ("Bélgica","Octavos de Final"),("Países Bajos","Octavos de Final"),
    ("Francia","Octavos de Final"),("Rumania","Octavos de Final"),
    ("Egipto","Octavos de Final"),("Argentina","Octavos de Final"),
],
1938:[
    ("Italia","Campeón"),("Hungría","Subcampeón"),("Brasil","Tercero"),
    ("Suecia","Cuarto"),("Francia","Cuartos de Final"),("Checoeslovaquia","Cuartos de Final"),
    ("Cuba","Cuartos de Final"),("Suiza","Cuartos de Final"),
    ("Noruega","Octavos de Final"),("Bélgica","Octavos de Final"),
    ("Alemania","Octavos de Final"),("Rumania","Octavos de Final"),
    ("Países Bajos","Octavos de Final"),("Polonia","Octavos de Final"),
    ("Indonesia","Octavos de Final"),  # Indias Orientales Neerlandesas
],
1950:[
    ("Uruguay","Campeón"),("Brasil","Subcampeón"),("Suecia","Tercero"),
    ("España","Cuarto"),("Bolivia","Fase de Grupos"),("Chile","Fase de Grupos"),
    ("Estados Unidos","Fase de Grupos"),("Inglaterra","Fase de Grupos"),
    ("Paraguay","Fase de Grupos"),("México","Fase de Grupos"),
    ("Yugoslavia","Fase de Grupos"),("Italia","Fase de Grupos"),("Suiza","Fase de Grupos"),
],
1954:[
    ("Alemania","Campeón"),("Hungría","Subcampeón"),("Austria","Tercero"),
    ("Uruguay","Cuarto"),("Yugoslavia","Cuartos de Final"),("Suiza","Cuartos de Final"),
    ("Brasil","Cuartos de Final"),("Inglaterra","Cuartos de Final"),
    ("Italia","Fase de Grupos"),("Bélgica","Fase de Grupos"),("Francia","Fase de Grupos"),
    ("México","Fase de Grupos"),("Checoeslovaquia","Fase de Grupos"),
    ("Escocia","Fase de Grupos"),("Corea del Sur","Fase de Grupos"),("Turquía","Fase de Grupos"),
],
1958:[
    ("Brasil","Campeón"),("Suecia","Subcampeón"),("Francia","Tercero"),
    ("Alemania","Cuarto"),("Yugoslavia","Cuartos de Final"),("Países Bajos","Cuartos de Final"),
    ("Gales","Cuartos de Final"),("Unión Soviética","Cuartos de Final"),
    ("Argentina","Fase de Grupos"),("Austria","Fase de Grupos"),("Checoeslovaquia","Fase de Grupos"),
    ("Hungría","Fase de Grupos"),("Inglaterra","Fase de Grupos"),("México","Fase de Grupos"),
    ("Irlanda del Norte","Fase de Grupos"),("Paraguay","Fase de Grupos"),
    ("Escocia","Fase de Grupos"),
],
1962:[
    ("Brasil","Campeón"),("Checoeslovaquia","Subcampeón"),("Chile","Tercero"),
    ("Yugoslavia","Cuarto"),("Inglaterra","Cuartos de Final"),("Hungría","Cuartos de Final"),
    ("Unión Soviética","Cuartos de Final"),("Alemania","Cuartos de Final"),
    ("Argentina","Fase de Grupos"),("Bulgaria","Fase de Grupos"),("Colombia","Fase de Grupos"),
    ("España","Fase de Grupos"),("Italia","Fase de Grupos"),("México","Fase de Grupos"),
    ("Suiza","Fase de Grupos"),("Uruguay","Fase de Grupos"),
],
1966:[
    ("Inglaterra","Campeón"),("Alemania","Subcampeón"),("Portugal","Tercero"),
    ("Unión Soviética","Cuarto"),("Hungría","Cuartos de Final"),("Argentina","Cuartos de Final"),
    ("Corea del Norte","Cuartos de Final"),("Uruguay","Cuartos de Final"),
    ("Brasil","Fase de Grupos"),("Bulgaria","Fase de Grupos"),("Chile","Fase de Grupos"),
    ("Francia","Fase de Grupos"),("Italia","Fase de Grupos"),("México","Fase de Grupos"),
    ("España","Fase de Grupos"),("Suiza","Fase de Grupos"),
],
1970:[
    ("Brasil","Campeón"),("Italia","Subcampeón"),("Alemania","Tercero"),
    ("Uruguay","Cuarto"),("México","Cuartos de Final"),("Unión Soviética","Cuartos de Final"),
    ("Perú","Cuartos de Final"),("Inglaterra","Cuartos de Final"),
    ("Bélgica","Fase de Grupos"),("Bulgaria","Fase de Grupos"),("Checoslovaquia","Fase de Grupos"),
    ("El Salvador","Fase de Grupos"),("Israel","Fase de Grupos"),("Marruecos","Fase de Grupos"),
    ("Rumania","Fase de Grupos"),("Suecia","Fase de Grupos"),
],
1974:[
    ("Alemania","Campeón"),("Países Bajos","Subcampeón"),("Polonia","Tercero"),
    ("Brasil","Cuarto"),("Argentina","Cuartos de Final"),("Alemania Democrática","Cuartos de Final"),
    ("Yugoslavia","Cuartos de Final"),("Suecia","Cuartos de Final"),
    ("Escocia","Fase de Grupos"),("Bulgaria","Fase de Grupos"),("Uruguay","Fase de Grupos"),
    ("Zaire","Fase de Grupos"),("Australia","Fase de Grupos"),("Chile","Fase de Grupos"),
    ("Italia","Fase de Grupos"),("Haití","Fase de Grupos"),
],
1978:[
    ("Argentina","Campeón"),("Países Bajos","Subcampeón"),("Brasil","Tercero"),
    ("Italia","Cuarto"),("Austria","Cuartos de Final"),("Alemania","Cuartos de Final"),
    ("Polonia","Cuartos de Final"),("Perú","Cuartos de Final"),
    ("México","Fase de Grupos"),("Francia","Fase de Grupos"),("Hungría","Fase de Grupos"),
    ("Irán","Fase de Grupos"),("Escocia","Fase de Grupos"),("Suecia","Fase de Grupos"),
    ("Túnez","Fase de Grupos"),("España","Fase de Grupos"),
],
1982:[
    ("Italia","Campeón"),("Alemania","Subcampeón"),("Polonia","Tercero"),
    ("Francia","Cuarto"),("España","Cuartos de Final"),("Bélgica","Cuartos de Final"),
    ("Austria","Cuartos de Final"),("Inglaterra","Cuartos de Final"),
    ("Argentina","Octavos de Final"),("Brasil","Octavos de Final"),("Hungría","Fase de Grupos"),
    ("Checoslovaquia","Fase de Grupos"),("Escocia","Fase de Grupos"),("Dinamarca","Fase de Grupos"),
    ("Países Bajos","Fase de Grupos"),("Kuwait","Fase de Grupos"),("Camerún","Fase de Grupos"),
    ("Honduras","Fase de Grupos"),("Yugoslavia","Fase de Grupos"),("Argelia","Fase de Grupos"),
    ("Perú","Fase de Grupos"),("Chile","Fase de Grupos"),("El Salvador","Fase de Grupos"),
    ("Nueva Zelanda","Fase de Grupos"),
],
1986:[
    ("Argentina","Campeón"),("Alemania","Subcampeón"),("Francia","Tercero"),
    ("Bélgica","Cuarto"),("España","Cuartos de Final"),("Inglaterra","Cuartos de Final"),
    ("Brasil","Cuartos de Final"),("México","Cuartos de Final"),
    ("Dinamarca","Octavos de Final"),("Marruecos","Octavos de Final"),
    ("Unión Soviética","Octavos de Final"),("Bulgaria","Octavos de Final"),
    ("Uruguay","Octavos de Final"),("Países Bajos","Octavos de Final"),
    ("Italia","Octavos de Final"),("Paraguay","Octavos de Final"),
    ("Iraq","Fase de Grupos"),("Portugal","Fase de Grupos"),("Corea del Sur","Fase de Grupos"),
    ("Canadá","Fase de Grupos"),("Hungría","Fase de Grupos"),("Argelia","Fase de Grupos"),
    ("Yugoslavia","Fase de Grupos"),("Escocia","Fase de Grupos"),
],
1990:[
    ("Alemania","Campeón"),("Argentina","Subcampeón"),("Italia","Tercero"),
    ("Inglaterra","Cuarto"),("Yugoslavia","Cuartos de Final"),("Checoeslovaquia","Cuartos de Final"),
    ("República de Irlanda","Cuartos de Final"),("Camerún","Cuartos de Final"),
    ("Brasil","Octavos de Final"),("Rumania","Octavos de Final"),
    ("Bélgica","Octavos de Final"),("Colombia","Octavos de Final"),
    ("España","Octavos de Final"),("Uruguay","Octavos de Final"),
    ("Costa Rica","Octavos de Final"),("Países Bajos","Octavos de Final"),
    ("Suecia","Fase de Grupos"),("Austria","Fase de Grupos"),("Escocia","Fase de Grupos"),
    ("Suiza","Fase de Grupos"),("Corea del Sur","Fase de Grupos"),("Ecuador","Fase de Grupos"),
    ("Egipto","Fase de Grupos"),("Estados Unidos","Fase de Grupos"),
],
1994:[
    ("Brasil","Campeón"),("Italia","Subcampeón"),("Suecia","Tercero"),
    ("Bulgaria","Cuarto"),("Rumania","Cuartos de Final"),("Países Bajos","Cuartos de Final"),
    ("Alemania","Cuartos de Final"),("España","Cuartos de Final"),
    ("Argentina","Octavos de Final"),("México","Octavos de Final"),
    ("Bélgica","Octavos de Final"),("Suiza","Octavos de Final"),
    ("Nigeria","Octavos de Final"),("Arabia Saudita","Octavos de Final"),
    ("Estados Unidos","Octavos de Final"),("República de Irlanda","Octavos de Final"),
    ("Bolivia","Fase de Grupos"),("Camerún","Fase de Grupos"),("Colombia","Fase de Grupos"),
    ("Corea del Sur","Fase de Grupos"),("Grecia","Fase de Grupos"),("Marruecos","Fase de Grupos"),
    ("Noruega","Fase de Grupos"),("Rusia","Fase de Grupos"),
],
1998:[
    ("Francia","Campeón"),("Brasil","Subcampeón"),("Croacia","Tercero"),
    ("Países Bajos","Cuarto"),("Alemania","Cuartos de Final"),("Argentina","Cuartos de Final"),
    ("Italia","Cuartos de Final"),("Dinamarca","Cuartos de Final"),
    ("México","Octavos de Final"),("Nigeria","Octavos de Final"),("Rumania","Octavos de Final"),
    ("Yugoslavia","Octavos de Final"),("Chile","Octavos de Final"),("Paraguay","Octavos de Final"),
    ("Inglaterra","Octavos de Final"),("Noruega","Octavos de Final"),
    ("Arabia Saudita","Fase de Grupos"),("Austria","Fase de Grupos"),
    ("Bélgica","Fase de Grupos"),("Bulgaria","Fase de Grupos"),("Camerún","Fase de Grupos"),
    ("Colombia","Fase de Grupos"),("Corea del Sur","Fase de Grupos"),("Ecuador","Fase de Grupos"),
    ("España","Fase de Grupos"),("Estados Unidos","Fase de Grupos"),("Jamaica","Fase de Grupos"),
    ("Japón","Fase de Grupos"),("Marruecos","Fase de Grupos"),("Sudáfrica","Fase de Grupos"),
    ("Túnez","Fase de Grupos"),("Irán","Fase de Grupos"),
],
2002:[
    ("Brasil","Campeón"),("Alemania","Subcampeón"),("Turquía","Tercero"),
    ("Corea del Sur","Cuarto"),("España","Cuartos de Final"),("Inglaterra","Cuartos de Final"),
    ("Senegal","Cuartos de Final"),("Estados Unidos","Cuartos de Final"),
    ("México","Octavos de Final"),("Paraguay","Octavos de Final"),("Japón","Octavos de Final"),
    ("Suecia","Octavos de Final"),("Bélgica","Octavos de Final"),("Dinamarca","Octavos de Final"),
    ("Irlanda","Octavos de Final"),("Ecuador","Octavos de Final"),
    ("Argentina","Fase de Grupos"),("Croacia","Fase de Grupos"),("Italia","Fase de Grupos"),
    ("Camerún","Fase de Grupos"),("Polonia","Fase de Grupos"),("Portugal","Fase de Grupos"),
    ("Arabia Saudita","Fase de Grupos"),("China","Fase de Grupos"),("Costa Rica","Fase de Grupos"),
    ("Nigeria","Fase de Grupos"),("Rusia","Fase de Grupos"),("Túnez","Fase de Grupos"),
    ("Eslovenia","Fase de Grupos"),("Uruguay","Fase de Grupos"),("Marruecos","Fase de Grupos"),
    ("Francia","Fase de Grupos"),
],
2006:[
    ("Italia","Campeón"),("Francia","Subcampeón"),("Alemania","Tercero"),
    ("Portugal","Cuarto"),("Argentina","Cuartos de Final"),("Inglaterra","Cuartos de Final"),
    ("Brasil","Cuartos de Final"),("España","Cuartos de Final"),
    ("México","Octavos de Final"),("Suiza","Octavos de Final"),("Australia","Octavos de Final"),
    ("Ecuador","Octavos de Final"),("Ucrania","Octavos de Final"),("Ghana","Octavos de Final"),
    ("Suecia","Octavos de Final"),("Japón","Octavos de Final"),
    ("Arabia Saudita","Fase de Grupos"),("Angola","Fase de Grupos"),("Corea del Sur","Fase de Grupos"),
    ("Costa Rica","Fase de Grupos"),("Croacia","Fase de Grupos"),("Eslovaquia","Fase de Grupos"),
    ("Estados Unidos","Fase de Grupos"),("Irán","Fase de Grupos"),("Países Bajos","Fase de Grupos"),
    ("Paraguay","Fase de Grupos"),("Polonia","Fase de Grupos"),("Serbia y Montenegro","Fase de Grupos"),
    ("Togo","Fase de Grupos"),("Trinidad y Tobago","Fase de Grupos"),
    ("Túnez","Fase de Grupos"),("Costa de Marfil","Fase de Grupos"),
],
2010:[
    ("España","Campeón"),("Países Bajos","Subcampeón"),("Alemania","Tercero"),
    ("Uruguay","Cuarto"),("Argentina","Cuartos de Final"),("Brasil","Cuartos de Final"),
    ("Ghana","Cuartos de Final"),("Paraguay","Cuartos de Final"),
    ("Corea del Sur","Octavos de Final"),("Estados Unidos","Octavos de Final"),
    ("México","Octavos de Final"),("Eslovaquia","Octavos de Final"),
    ("Chile","Octavos de Final"),("Portugal","Octavos de Final"),
    ("Inglaterra","Octavos de Final"),("Japón","Octavos de Final"),
    ("Argelia","Fase de Grupos"),("Australia","Fase de Grupos"),("Camerún","Fase de Grupos"),
    ("Costa de Marfil","Fase de Grupos"),("Costa Rica","Fase de Grupos"),
    ("Dinamarca","Fase de Grupos"),("Francia","Fase de Grupos"),("Grecia","Fase de Grupos"),
    ("Honduras","Fase de Grupos"),("Italia","Fase de Grupos"),("Nigeria","Fase de Grupos"),
    ("Nueva Zelanda","Fase de Grupos"),("Serbia","Fase de Grupos"),
    ("Eslovenia","Fase de Grupos"),("Sudáfrica","Fase de Grupos"),("Suiza","Fase de Grupos"),
],
2014:[
    ("Alemania","Campeón"),("Argentina","Subcampeón"),("Países Bajos","Tercero"),
    ("Brasil","Cuarto"),("Francia","Cuartos de Final"),("Colombia","Cuartos de Final"),
    ("Costa Rica","Cuartos de Final"),("Bélgica","Cuartos de Final"),
    ("Suiza","Octavos de Final"),("México","Octavos de Final"),("Grecia","Octavos de Final"),
    ("Estados Unidos","Octavos de Final"),("Chile","Octavos de Final"),("Uruguay","Octavos de Final"),
    ("Nigeria","Octavos de Final"),("Argelia","Octavos de Final"),
    ("Australia","Fase de Grupos"),("Bosnia y Herzegovina","Fase de Grupos"),
    ("Camerún","Fase de Grupos"),("Costa de Marfil","Fase de Grupos"),
    ("Croacia","Fase de Grupos"),("Ecuador","Fase de Grupos"),("Ghana","Fase de Grupos"),
    ("Honduras","Fase de Grupos"),("Irán","Fase de Grupos"),("Italia","Fase de Grupos"),
    ("Japón","Fase de Grupos"),("Corea del Sur","Fase de Grupos"),
    ("Portugal","Fase de Grupos"),("Rusia","Fase de Grupos"),("España","Fase de Grupos"),
    ("Inglaterra","Fase de Grupos"),
],
2018:[
    ("Francia","Campeón"),("Croacia","Subcampeón"),("Bélgica","Tercero"),
    ("Inglaterra","Cuarto"),("Uruguay","Cuartos de Final"),("Brasil","Cuartos de Final"),
    ("Rusia","Cuartos de Final"),("Suecia","Cuartos de Final"),
    ("Argentina","Octavos de Final"),("España","Octavos de Final"),
    ("Dinamarca","Octavos de Final"),("Colombia","Octavos de Final"),
    ("Suiza","Octavos de Final"),("Japón","Octavos de Final"),
    ("México","Octavos de Final"),("Portugal","Octavos de Final"),
    ("Alemania","Fase de Grupos"),("Arabia Saudita","Fase de Grupos"),
    ("Australia","Fase de Grupos"),("Corea del Sur","Fase de Grupos"),
    ("Costa Rica","Fase de Grupos"),("Egipto","Fase de Grupos"),
    ("Irán","Fase de Grupos"),("Islandia","Fase de Grupos"),
    ("Marruecos","Fase de Grupos"),("Nigeria","Fase de Grupos"),
    ("Panamá","Fase de Grupos"),("Perú","Fase de Grupos"),
    ("Polonia","Fase de Grupos"),("Senegal","Fase de Grupos"),
    ("Serbia","Fase de Grupos"),("Túnez","Fase de Grupos"),
],
2022:[
    ("Argentina","Campeón"),("Francia","Subcampeón"),("Croacia","Tercero"),
    ("Marruecos","Cuarto"),("Países Bajos","Cuartos de Final"),("Portugal","Cuartos de Final"),
    ("Inglaterra","Cuartos de Final"),("Brasil","Cuartos de Final"),
    ("Japón","Octavos de Final"),("Senegal","Octavos de Final"),
    ("España","Octavos de Final"),("Estados Unidos","Octavos de Final"),
    ("Australia","Octavos de Final"),("Suiza","Octavos de Final"),
    ("Corea del Sur","Octavos de Final"),("Polonia","Octavos de Final"),
    ("Alemania","Fase de Grupos"),("Arabia Saudita","Fase de Grupos"),
    ("Bélgica","Fase de Grupos"),("Camerún","Fase de Grupos"),
    ("Canadá","Fase de Grupos"),("Costa Rica","Fase de Grupos"),
    ("Dinamarca","Fase de Grupos"),("Ecuador","Fase de Grupos"),
    ("Gales","Fase de Grupos"),("Ghana","Fase de Grupos"),
    ("Irán","Fase de Grupos"),("México","Fase de Grupos"),
    ("Serbia","Fase de Grupos"),("Túnez","Fase de Grupos"),
    ("Catar","Fase de Grupos"),("Uruguay","Fase de Grupos"),
],
}

SEDE_ISO = {
    "Uruguay":"URY","Italia":"ITA","Francia":"FRA","Brasil":"BRA",
    "Suiza":"CHE","Suecia":"SWE","Chile":"CHL","Inglaterra":"GBR",
    "México":"MEX","Alemania":"DEU","Argentina":"ARG","España":"ESP",
    "Estados Unidos":"USA","Rusia":"RUS","Sudáfrica":"ZAF","Catar":"QAT",
    "Corea/Japón":"KOR",
}

# ─────────────────────────────────────────────────────────────────────────────
# 2.  CONSTRUIR DATAFRAMES
# ─────────────────────────────────────────────────────────────────────────────

def conf_row(pais):
    if pais in CONF:
        return CONF[pais]
    # fallback búsqueda parcial
    for k,v in CONF.items():
        if k.lower() in pais.lower() or pais.lower() in k.lower():
            return v
    return ("?","Desconocido","???")

# --- EDICIONES ---
ediciones_rows = []
for (ano,sede,n,camp,sub,ter,cua) in EDICIONES_RAW:
    cf_camp = conf_row(camp)[0]
    cf_sub  = conf_row(sub)[0]
    cf_ter  = conf_row(ter)[0]
    cf_cua  = conf_row(cua)[0]
    ediciones_rows.append({
        "Año":ano, "Sede":sede, "N_Equipos":n,
        "Campeón":camp,     "Conf_Campeón":cf_camp,
        "Subcampeón":sub,   "Conf_Subcampeón":cf_sub,
        "Tercero":ter,      "Conf_Tercero":cf_ter,
        "Cuarto":cua,       "Conf_Cuarto":cf_cua,
    })
df_ediciones = pd.DataFrame(ediciones_rows)

# --- PARTICIPACIONES FLAT ---
part_rows = []
for ano, lista in PARTICIPACIONES_RAW.items():
    for (pais, fase) in lista:
        cr = conf_row(pais)
        part_rows.append({
            "Año":ano,
            "País":pais,
            "ISO_Code":cr[2],
            "Confederación":cr[0],
            "Confederación_Nombre":cr[1],
            "Fase":fase,
            "Fase_Num":FASE_NUM.get(fase,0),
            "Es_Campeón":   1 if fase=="Campeón" else 0,
            "Es_Subcampeón":1 if fase=="Subcampeón" else 0,
            "Es_Tercero":   1 if fase=="Tercero" else 0,
            "Es_Cuarto":    1 if fase=="Cuarto" else 0,
            "En_Final":     1 if fase in ("Campeón","Subcampeón") else 0,
            "En_Podio":     1 if fase in ("Campeón","Subcampeón","Tercero") else 0,
            "En_Top4":      1 if fase in ("Campeón","Subcampeón","Tercero","Cuarto") else 0,
        })
df_part = pd.DataFrame(part_rows)

# --- RESUMEN POR CONFEDERACIÓN ---
df_rconf = df_part.groupby("Confederación").agg(
    Nombre         =("Confederación_Nombre","first"),
    N_Participaciones=("Año","count"),
    N_Mundiales    =("Año","nunique"),
    Campeones      =("Es_Campeón","sum"),
    Subcampeones   =("Es_Subcampeón","sum"),
    Terceros       =("Es_Tercero","sum"),
    Cuartos        =("Es_Cuarto","sum"),
    Finales        =("En_Final","sum"),
    Podios         =("En_Podio","sum"),
    Top4           =("En_Top4","sum"),
).reset_index()
df_rconf["Medallas_Total"] = df_rconf["Campeones"]+df_rconf["Subcampeones"]+df_rconf["Terceros"]
df_rconf = df_rconf.sort_values("Campeones",ascending=False)

# --- RESUMEN POR PAÍS ---
df_rpais = df_part.groupby(["País","ISO_Code","Confederación","Confederación_Nombre"]).agg(
    N_Mundiales  =("Año","nunique"),
    Mejor_Fase   =("Fase_Num","max"),
    Campeones    =("Es_Campeón","sum"),
    Subcampeones =("Es_Subcampeón","sum"),
    Terceros     =("Es_Tercero","sum"),
    Cuartos      =("Es_Cuarto","sum"),
    Finales      =("En_Final","sum"),
    Top4         =("En_Top4","sum"),
).reset_index()
FASE_LBL={6:"Campeón",5:"Subcampeón",4:"Tercero",3:"Cuarto",2:"Cuartos de Final",1:"Octavos de Final",0:"Fase de Grupos"}
df_rpais["Mejor_Fase_Nombre"]=df_rpais["Mejor_Fase"].map(FASE_LBL)
df_rpais = df_rpais.sort_values(["Campeones","N_Mundiales"],ascending=[False,False])

# ─────────────────────────────────────────────────────────────────────────────
# 3.  CREAR EXCEL
# ─────────────────────────────────────────────────────────────────────────────

AMARILLO = "FFFF00"
DORADO   = "FFD700"
GRIS_H   = "D9D9D9"
AZUL_H   = "1F3864"
NEGRO    = "000000"
BLANCO   = "FFFFFF"

def hdr_style(cell, bg=AZUL_H, fg=BLANCO, bold=True):
    cell.font      = Font(bold=bold, color=fg, name="Arial", size=11)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=NEGRO)
    cell.border = Border(thin,thin,thin,thin)

def data_style(cell, bg=None, bold=False, center=False):
    cell.font      = Font(bold=bold, color=NEGRO, name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    cell.border = Border(thin,thin,thin,thin)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

def auto_width(ws, extra=4):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+extra, 40)

def write_df(ws, df, start_row=2, header_bg=AZUL_H):
    # Header row 1
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=start_row, column=ci, value=col)
        hdr_style(c, bg=header_bg)
    # Data rows
    alt = ["F2F2F2", BLANCO]
    for ri, (_, row) in enumerate(df.iterrows()):
        bg = alt[ri % 2]
        for ci, col in enumerate(df.columns, 1):
            val = row[col]
            c = ws.cell(row=start_row+1+ri, column=ci, value=val)
            data_style(c, bg=bg, center=True)
    ws.row_dimensions[start_row].height = 30
    ws.freeze_panes = f"A{start_row+1}"

wb = openpyxl.Workbook()

# ── Sheet 1: Ediciones ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Ediciones"
ws1.sheet_view.showGridLines = False
write_df(ws1, df_ediciones, start_row=1)
auto_width(ws1)

# ── Sheet 2: Participaciones ─────────────────────────────────────────────────
ws2 = wb.create_sheet("Participaciones")
ws2.sheet_view.showGridLines = False
write_df(ws2, df_part, start_row=1)
auto_width(ws2)

# ── Sheet 3: Resumen Confederación ───────────────────────────────────────────
ws3 = wb.create_sheet("Resumen_Confederación")
ws3.sheet_view.showGridLines = False
write_df(ws3, df_rconf, start_row=1)
auto_width(ws3)

# ── Sheet 4: Resumen País ────────────────────────────────────────────────────
ws4 = wb.create_sheet("Resumen_País")
ws4.sheet_view.showGridLines = False
write_df(ws4, df_rpais, start_row=1)
auto_width(ws4)

# ── Sheet 5: Datos_Mapa ──────────────────────────────────────────────────────
ws5 = wb.create_sheet("Datos_Mapa")
ws5.sheet_view.showGridLines = False
df_mapa = df_rpais[["País","ISO_Code","Confederación","Confederación_Nombre",
                     "N_Mundiales","Campeones","Mejor_Fase_Nombre"]].copy()
df_mapa.rename(columns={"Confederación":"Conf_Cod","Confederación_Nombre":"Conf_Nombre"},inplace=True)
write_df(ws5, df_mapa, start_row=1)
auto_width(ws5)

excel_path = os.path.join(OUT_DIR, EXCEL_NAME)
wb.save(excel_path)
print(f"[OK] Excel guardado: {excel_path}")

# ─────────────────────────────────────────────────────────────────────────────
# 4.  CREAR TABLEAU WORKBOOK (.twb)
# ─────────────────────────────────────────────────────────────────────────────

# Colores por confederación
CONF_COLORS = {
    "UEFA":     "#1f77b4",  # azul
    "CONMEBOL": "#2ca02c",  # verde
    "CAF":      "#d62728",  # rojo
    "AFC":      "#ff7f0e",  # naranja
    "CONCACAF": "#9467bd",  # violeta
    "OFC":      "#17becf",  # cyan
}

twb_xml = """\
<?xml version='1.0' encoding='utf-8' ?>

<!-- =========================================================
     Mundiales por Confederaciones  |  Tableau Workbook
     Creado con Python  |  Datos: FIFA / Wikipedia
     ========================================================= -->

<workbook source-build='2023.3.0.1002' source-platform='win' version='18.1'
          xmlns:user='http://www.tableausoftware.com/xml/user'>

  <!-- ── PREFERENCIAS ──────────────────────────────────── -->
  <preferences>
    <color-palette name='Confederaciones' type='regular'>
      <color>#1f77b4</color>  <!-- UEFA   -->
      <color>#2ca02c</color>  <!-- CONMEBOL -->
      <color>#d62728</color>  <!-- CAF    -->
      <color>#ff7f0e</color>  <!-- AFC    -->
      <color>#9467bd</color>  <!-- CONCACAF -->
      <color>#17becf</color>  <!-- OFC   -->
    </color-palette>
  </preferences>

  <!-- ── FUENTES DE DATOS ──────────────────────────────── -->
  <datasources>

    <!-- 1.  Participaciones (tabla principal) -->
    <datasource caption='Mundiales - Participaciones' inline='true'
                name='mw-participaciones' version='18.1'>
      <connection class='excel-direct' cleaning='no' compat='no'
                  dataRefreshTime='' filename='./Data/Mundiales_Datos.xlsx'
                  interpret_numbers_as_text='no' validate='no'>
        <relation name='Participaciones$' table='[Participaciones$]' type='table'/>
      </connection>
      <aliases enabled='yes'/>
      <column caption='Año'               datatype='integer' name='[Año]'               role='dimension' type='ordinal'/>
      <column caption='País'              datatype='string'  name='[País]'              role='dimension' type='nominal'/>
      <column caption='ISO Code'          datatype='string'  name='[ISO_Code]'          role='dimension' type='nominal'>
        <geographic-role>iso-3166-1-alpha-3</geographic-role>
      </column>
      <column caption='Confederación'     datatype='string'  name='[Confederación]'     role='dimension' type='nominal'/>
      <column caption='Conf. Nombre'      datatype='string'  name='[Confederación_Nombre]' role='dimension' type='nominal'/>
      <column caption='Fase'              datatype='string'  name='[Fase]'              role='dimension' type='nominal'/>
      <column caption='Fase Num'          datatype='integer' name='[Fase_Num]'          role='measure'   type='quantitative'/>
      <column caption='Es Campeón'        datatype='integer' name='[Es_Campeón]'        role='measure'   type='quantitative'/>
      <column caption='Es Subcampeón'     datatype='integer' name='[Es_Subcampeón]'     role='measure'   type='quantitative'/>
      <column caption='Es Tercero'        datatype='integer' name='[Es_Tercero]'        role='measure'   type='quantitative'/>
      <column caption='Es Cuarto'         datatype='integer' name='[Es_Cuarto]'         role='measure'   type='quantitative'/>
      <column caption='En Final'          datatype='integer' name='[En_Final]'          role='measure'   type='quantitative'/>
      <column caption='En Podio'          datatype='integer' name='[En_Podio]'          role='measure'   type='quantitative'/>
      <column caption='En Top 4'          datatype='integer' name='[En_Top4]'           role='measure'   type='quantitative'/>
    </datasource>

    <!-- 2.  Ediciones -->
    <datasource caption='Mundiales - Ediciones' inline='true'
                name='mw-ediciones' version='18.1'>
      <connection class='excel-direct' cleaning='no' compat='no'
                  dataRefreshTime='' filename='./Data/Mundiales_Datos.xlsx'
                  interpret_numbers_as_text='no' validate='no'>
        <relation name='Ediciones$' table='[Ediciones$]' type='table'/>
      </connection>
      <aliases enabled='yes'/>
      <column caption='Año'               datatype='integer' name='[Año]'               role='dimension' type='ordinal'/>
      <column caption='Sede'              datatype='string'  name='[Sede]'              role='dimension' type='nominal'/>
      <column caption='N Equipos'         datatype='integer' name='[N_Equipos]'         role='measure'   type='quantitative'/>
      <column caption='Campeón'           datatype='string'  name='[Campeón]'           role='dimension' type='nominal'/>
      <column caption='Conf. Campeón'     datatype='string'  name='[Conf_Campeón]'      role='dimension' type='nominal'/>
      <column caption='Subcampeón'        datatype='string'  name='[Subcampeón]'        role='dimension' type='nominal'/>
      <column caption='Conf. Subcampeón'  datatype='string'  name='[Conf_Subcampeón]'   role='dimension' type='nominal'/>
      <column caption='Tercero'           datatype='string'  name='[Tercero]'           role='dimension' type='nominal'/>
      <column caption='Conf. Tercero'     datatype='string'  name='[Conf_Tercero]'      role='dimension' type='nominal'/>
      <column caption='Cuarto'            datatype='string'  name='[Cuarto]'            role='dimension' type='nominal'/>
      <column caption='Conf. Cuarto'      datatype='string'  name='[Conf_Cuarto]'       role='dimension' type='nominal'/>
    </datasource>

    <!-- 3.  Resumen por Confederación -->
    <datasource caption='Mundiales - Confederaciones' inline='true'
                name='mw-confederaciones' version='18.1'>
      <connection class='excel-direct' cleaning='no' compat='no'
                  dataRefreshTime='' filename='./Data/Mundiales_Datos.xlsx'
                  interpret_numbers_as_text='no' validate='no'>
        <relation name='Resumen_Confederación$' table='[Resumen_Confederación$]' type='table'/>
      </connection>
      <aliases enabled='yes'/>
      <column caption='Confederación'     datatype='string'  name='[Confederación]'     role='dimension' type='nominal'/>
      <column caption='Nombre'            datatype='string'  name='[Nombre]'            role='dimension' type='nominal'/>
      <column caption='N Participaciones' datatype='integer' name='[N_Participaciones]' role='measure'   type='quantitative'/>
      <column caption='N Mundiales'       datatype='integer' name='[N_Mundiales]'       role='measure'   type='quantitative'/>
      <column caption='Campeones'         datatype='integer' name='[Campeones]'         role='measure'   type='quantitative'/>
      <column caption='Subcampeones'      datatype='integer' name='[Subcampeones]'      role='measure'   type='quantitative'/>
      <column caption='Terceros'          datatype='integer' name='[Terceros]'          role='measure'   type='quantitative'/>
      <column caption='Cuartos'           datatype='integer' name='[Cuartos]'           role='measure'   type='quantitative'/>
      <column caption='Finales'           datatype='integer' name='[Finales]'           role='measure'   type='quantitative'/>
      <column caption='Podios'            datatype='integer' name='[Podios]'            role='measure'   type='quantitative'/>
      <column caption='Top 4'             datatype='integer' name='[Top4]'              role='measure'   type='quantitative'/>
      <column caption='Medallas Total'    datatype='integer' name='[Medallas_Total]'    role='measure'   type='quantitative'/>
    </datasource>

    <!-- 4.  Resumen por País -->
    <datasource caption='Mundiales - Países' inline='true'
                name='mw-paises' version='18.1'>
      <connection class='excel-direct' cleaning='no' compat='no'
                  dataRefreshTime='' filename='./Data/Mundiales_Datos.xlsx'
                  interpret_numbers_as_text='no' validate='no'>
        <relation name='Resumen_País$' table='[Resumen_País$]' type='table'/>
      </connection>
      <aliases enabled='yes'/>
      <column caption='País'              datatype='string'  name='[País]'              role='dimension' type='nominal'/>
      <column caption='ISO Code'          datatype='string'  name='[ISO_Code]'          role='dimension' type='nominal'>
        <geographic-role>iso-3166-1-alpha-3</geographic-role>
      </column>
      <column caption='Confederación'     datatype='string'  name='[Confederación]'     role='dimension' type='nominal'/>
      <column caption='N Mundiales'       datatype='integer' name='[N_Mundiales]'       role='measure'   type='quantitative'/>
      <column caption='Mejor Fase'        datatype='integer' name='[Mejor_Fase]'        role='measure'   type='quantitative'/>
      <column caption='Mejor Fase Nombre' datatype='string'  name='[Mejor_Fase_Nombre]' role='dimension' type='nominal'/>
      <column caption='Campeones'         datatype='integer' name='[Campeones]'         role='measure'   type='quantitative'/>
      <column caption='Subcampeones'      datatype='integer' name='[Subcampeones]'      role='measure'   type='quantitative'/>
      <column caption='Terceros'          datatype='integer' name='[Terceros]'          role='measure'   type='quantitative'/>
      <column caption='Finales'           datatype='integer' name='[Finales]'           role='measure'   type='quantitative'/>
      <column caption='Top 4'             datatype='integer' name='[Top4]'              role='measure'   type='quantitative'/>
    </datasource>

  </datasources>

  <!-- ── HOJAS DE TRABAJO ──────────────────────────────── -->
  <worksheets>

    <!-- ① Campeonatos por Confederación (barras horizontales) -->
    <worksheet name='Campeonatos x Confederación'>
      <table>
        <view>
          <datasources>
            <datasource caption='Mundiales - Confederaciones' name='mw-confederaciones'/>
          </datasources>
          <datasource-dependencies datasource='mw-confederaciones'>
            <column name='[Confederación]'/>
            <column name='[Campeones]'/>
          </datasource-dependencies>
        </view>
        <style>
          <style-rule element='mark'>
            <encoding attr='color' field='[mw-confederaciones].[Confederación]'
                      palette='Confederaciones' type='palette'/>
          </style-rule>
          <style-rule element='header'>
            <encoding attr='font-size'   value='12'/>
          </style-rule>
          <style-rule element='axis'>
            <encoding attr='font-size'   value='11'/>
          </style-rule>
        </style>
        <panes>
          <pane>
            <view>
              <mark class='Bar'/>
            </view>
            <encodings>
              <color field='[mw-confederaciones].[Confederación]'
                     palette='Confederaciones' type='palette'/>
              <label field='[mw-confederaciones].[Campeones]'/>
            </encodings>
          </pane>
        </panes>
        <rows>[mw-confederaciones].[Confederación]</rows>
        <cols>SUM([mw-confederaciones].[Campeones])</cols>
        <aliases>
          <alias key='UEFA'     value='UEFA (Europa)'/>
          <alias key='CONMEBOL' value='CONMEBOL (Am. Sur)'/>
          <alias key='CAF'      value='CAF (África)'/>
          <alias key='AFC'      value='AFC (Asia)'/>
          <alias key='CONCACAF' value='CONCACAF (Am. Norte)'/>
          <alias key='OFC'      value='OFC (Oceanía)'/>
        </aliases>
      </table>
    </worksheet>

    <!-- ② Participaciones por Confederación -->
    <worksheet name='Participaciones x Confederación'>
      <table>
        <view>
          <datasources>
            <datasource caption='Mundiales - Confederaciones' name='mw-confederaciones'/>
          </datasources>
          <datasource-dependencies datasource='mw-confederaciones'>
            <column name='[Confederación]'/>
            <column name='[N_Participaciones]'/>
          </datasource-dependencies>
        </view>
        <style>
          <style-rule element='mark'>
            <encoding attr='color' field='[mw-confederaciones].[Confederación]'
                      palette='Confederaciones' type='palette'/>
          </style-rule>
        </style>
        <panes>
          <pane>
            <view>
              <mark class='Bar'/>
            </view>
            <encodings>
              <color field='[mw-confederaciones].[Confederación]'
                     palette='Confederaciones' type='palette'/>
              <label field='[mw-confederaciones].[N_Participaciones]'/>
            </encodings>
          </pane>
        </panes>
        <rows>[mw-confederaciones].[Confederación]</rows>
        <cols>SUM([mw-confederaciones].[N_Participaciones])</cols>
      </table>
    </worksheet>

    <!-- ③ Medallero por Confederación (barras apiladas) -->
    <worksheet name='Medallero x Confederación'>
      <table>
        <view>
          <datasources>
            <datasource caption='Mundiales - Confederaciones' name='mw-confederaciones'/>
          </datasources>
          <datasource-dependencies datasource='mw-confederaciones'>
            <column name='[Confederación]'/>
            <column name='[Campeones]'/>
            <column name='[Subcampeones]'/>
            <column name='[Terceros]'/>
          </datasource-dependencies>
        </view>
        <style>
          <style-rule element='mark'>
            <encoding attr='color' field='[mw-confederaciones].[Confederación]'
                      palette='Confederaciones' type='palette'/>
          </style-rule>
        </style>
        <panes>
          <pane>
            <view>
              <mark class='Bar'/>
            </view>
            <encodings>
              <color field='[mw-confederaciones].[Confederación]'
                     palette='Confederaciones' type='palette'/>
            </encodings>
          </pane>
        </panes>
        <rows>[mw-confederaciones].[Confederación]</rows>
        <cols>SUM([mw-confederaciones].[Medallas_Total])</cols>
      </table>
    </worksheet>

    <!-- ④ Top Países por Apariciones -->
    <worksheet name='Top Países - Apariciones'>
      <table>
        <view>
          <datasources>
            <datasource caption='Mundiales - Países' name='mw-paises'/>
          </datasources>
          <datasource-dependencies datasource='mw-paises'>
            <column name='[País]'/>
            <column name='[N_Mundiales]'/>
            <column name='[Confederación]'/>
          </datasource-dependencies>
        </view>
        <style>
          <style-rule element='mark'>
            <encoding attr='color' field='[mw-paises].[Confederación]'
                      palette='Confederaciones' type='palette'/>
          </style-rule>
        </style>
        <panes>
          <pane>
            <view>
              <mark class='Bar'/>
            </view>
            <encodings>
              <color field='[mw-paises].[Confederación]'
                     palette='Confederaciones' type='palette'/>
              <label field='[mw-paises].[N_Mundiales]'/>
            </encodings>
          </pane>
        </panes>
        <rows>[mw-paises].[País]</rows>
        <cols>SUM([mw-paises].[N_Mundiales])</cols>
      </table>
    </worksheet>

    <!-- ⑤ Mapa Mundial de Participaciones -->
    <worksheet name='Mapa - Participaciones'>
      <table>
        <view>
          <datasources>
            <datasource caption='Mundiales - Países' name='mw-paises'/>
          </datasources>
          <datasource-dependencies datasource='mw-paises'>
            <column name='[ISO_Code]'/>
            <column name='[N_Mundiales]'/>
            <column name='[Confederación]'/>
            <column name='[País]'/>
          </datasource-dependencies>
        </view>
        <style>
          <style-rule element='mark'>
            <encoding attr='color' field='[mw-paises].[Confederación]'
                      palette='Confederaciones' type='palette'/>
          </style-rule>
        </style>
        <panes>
          <pane>
            <view>
              <mark class='Map'/>
            </view>
            <encodings>
              <color field='[mw-paises].[Confederación]'
                     palette='Confederaciones' type='palette'/>
              <size  field='[mw-paises].[N_Mundiales]'/>
              <label field='[mw-paises].[País]'/>
            </encodings>
          </pane>
        </panes>
        <rows>AVG([mw-paises].[Latitude (generated)])</rows>
        <cols>AVG([mw-paises].[Longitude (generated)])</cols>
      </table>
    </worksheet>

    <!-- ⑥ Historial – tabla de ediciones -->
    <worksheet name='Historial Ediciones'>
      <table>
        <view>
          <datasources>
            <datasource caption='Mundiales - Ediciones' name='mw-ediciones'/>
          </datasources>
          <datasource-dependencies datasource='mw-ediciones'>
            <column name='[Año]'/>
            <column name='[Sede]'/>
            <column name='[N_Equipos]'/>
            <column name='[Campeón]'/>
            <column name='[Conf_Campeón]'/>
            <column name='[Subcampeón]'/>
            <column name='[Tercero]'/>
            <column name='[Cuarto]'/>
          </datasource-dependencies>
        </view>
        <style>
          <style-rule element='mark'>
            <encoding attr='color' field='[mw-ediciones].[Conf_Campeón]'
                      palette='Confederaciones' type='palette'/>
          </style-rule>
        </style>
        <panes>
          <pane>
            <view>
              <mark class='Text'/>
            </view>
            <encodings>
              <text field='[mw-ediciones].[Campeón]'/>
              <color field='[mw-ediciones].[Conf_Campeón]'
                     palette='Confederaciones' type='palette'/>
            </encodings>
          </pane>
        </panes>
        <rows>[mw-ediciones].[Año]</rows>
        <cols>[mw-ediciones].[Sede]
              [mw-ediciones].[N_Equipos]
              [mw-ediciones].[Campeón]
              [mw-ediciones].[Subcampeón]
              [mw-ediciones].[Tercero]
              [mw-ediciones].[Cuarto]</cols>
      </table>
    </worksheet>

    <!-- ⑦ Evolución Histórica de participaciones por conf. -->
    <worksheet name='Evolución Histórica'>
      <table>
        <view>
          <datasources>
            <datasource caption='Mundiales - Participaciones' name='mw-participaciones'/>
          </datasources>
          <datasource-dependencies datasource='mw-participaciones'>
            <column name='[Año]'/>
            <column name='[Confederación]'/>
          </datasource-dependencies>
        </view>
        <style>
          <style-rule element='mark'>
            <encoding attr='color' field='[mw-participaciones].[Confederación]'
                      palette='Confederaciones' type='palette'/>
          </style-rule>
        </style>
        <panes>
          <pane>
            <view>
              <mark class='Bar'/>
            </view>
            <encodings>
              <color field='[mw-participaciones].[Confederación]'
                     palette='Confederaciones' type='palette'/>
            </encodings>
          </pane>
        </panes>
        <rows>CNT([mw-participaciones].[País])</rows>
        <cols>[mw-participaciones].[Año]</cols>
      </table>
    </worksheet>

    <!-- ⑧ KPIs – Texto grande por confederación -->
    <worksheet name='KPIs Confederaciones'>
      <table>
        <view>
          <datasources>
            <datasource caption='Mundiales - Confederaciones' name='mw-confederaciones'/>
          </datasources>
          <datasource-dependencies datasource='mw-confederaciones'>
            <column name='[Confederación]'/>
            <column name='[Campeones]'/>
            <column name='[N_Participaciones]'/>
            <column name='[Medallas_Total]'/>
          </datasource-dependencies>
        </view>
        <style>
          <style-rule element='mark'>
            <encoding attr='color' field='[mw-confederaciones].[Confederación]'
                      palette='Confederaciones' type='palette'/>
          </style-rule>
        </style>
        <panes>
          <pane>
            <view>
              <mark class='Text'/>
            </view>
            <encodings>
              <color field='[mw-confederaciones].[Confederación]'
                     palette='Confederaciones' type='palette'/>
              <text field='[mw-confederaciones].[Campeones]'/>
              <size field='[mw-confederaciones].[Campeones]'/>
            </encodings>
          </pane>
        </panes>
        <rows>[mw-confederaciones].[Confederación]</rows>
        <cols>SUM([mw-confederaciones].[Campeones])
              SUM([mw-confederaciones].[N_Participaciones])
              SUM([mw-confederaciones].[Medallas_Total])</cols>
      </table>
    </worksheet>

  </worksheets>

  <!-- ── DASHBOARD ─────────────────────────────────────── -->
  <dashboards>
    <dashboard name='🏆 Mundiales por Confederaciones'>
      <style>
        <style-rule element='dashboard'>
          <encoding attr='font-family' value='Arial'/>
          <encoding attr='font-size'   value='12'/>
        </style-rule>
      </style>
      <size maxheight='900' maxwidth='1600' minheight='900' minwidth='1600'/>
      <zones>
        <!-- Zona principal: layout vertical -->
        <zone h='900' id='1' type='layout-basic' w='1600' x='0' y='0'>
          <zone h='70' id='2' type='layout-basic' w='1600' x='0' y='0'>
            <!-- Texto de título superior (banner amarillo) -->
            <zone fixed-size='70' h='70' id='3' type='text' w='1600' x='0' y='0'>
              <formatted-text>
                <run bold='true' fontcolor='#000000' fontname='Arial Black'
                     fontsize='20' fontstyle='bold'>
                  🏆  MUNDIALES DE FÚTBOL POR CONFEDERACIÓN  |  1930 – 2022
                </run>
              </formatted-text>
              <zone-style>
                <format attr='background' value='#FFFF00'/>
                <format attr='text-align' value='center'/>
                <format attr='font-size'  value='20'/>
              </zone-style>
            </zone>
          </zone>

          <!-- Fila 2: tres columnas -->
          <zone h='410' id='10' type='layout-basic' w='1600' x='0' y='70'>

            <!-- Columna izq: KPIs + Mapa -->
            <zone h='410' id='11' type='layout-basic' w='430' x='0' y='70'>
              <zone h='200' id='12' name='KPIs Confederaciones' type='worksheet' w='430' x='0' y='70'>
                <zone-style><format attr='border' value='1'/></zone-style>
              </zone>
              <zone h='210' id='13' name='Campeonatos x Confederación' type='worksheet' w='430' x='0' y='270'>
                <zone-style><format attr='border' value='1'/></zone-style>
              </zone>
            </zone>

            <!-- Columna central: Mapa -->
            <zone h='410' id='14' name='Mapa - Participaciones' type='worksheet' w='740' x='430' y='70'>
              <zone-style><format attr='border' value='1'/></zone-style>
            </zone>

            <!-- Columna der: Participaciones + Medallero -->
            <zone h='410' id='15' type='layout-basic' w='430' x='1170' y='70'>
              <zone h='205' id='16' name='Participaciones x Confederación' type='worksheet' w='430' x='1170' y='70'>
                <zone-style><format attr='border' value='1'/></zone-style>
              </zone>
              <zone h='205' id='17' name='Medallero x Confederación' type='worksheet' w='430' x='1170' y='275'>
                <zone-style><format attr='border' value='1'/></zone-style>
              </zone>
            </zone>
          </zone>

          <!-- Fila 3: dos paneles inferiores -->
          <zone h='420' id='20' type='layout-basic' w='1600' x='0' y='480'>

            <!-- Panel inferior izq: Top Países -->
            <zone h='420' id='21' name='Top Países - Apariciones' type='worksheet' w='530' x='0' y='480'>
              <zone-style><format attr='border' value='1'/></zone-style>
            </zone>

            <!-- Panel inferior central: Historial -->
            <zone h='420' id='22' name='Historial Ediciones' type='worksheet' w='530' x='530' y='480'>
              <zone-style><format attr='border' value='1'/></zone-style>
            </zone>

            <!-- Panel inferior der: Evolución -->
            <zone h='420' id='23' name='Evolución Histórica' type='worksheet' w='540' x='1060' y='480'>
              <zone-style><format attr='border' value='1'/></zone-style>
            </zone>
          </zone>

        </zone>
      </zones>
      <devicelayouts>
        <devicelayout auto-generated='true' name='Phone'>
          <size maxheight='700' maxwidth='500' minheight='700' minwidth='500'/>
        </devicelayout>
      </devicelayouts>
    </dashboard>
  </dashboards>

</workbook>
"""

twb_path = os.path.join(OUT_DIR, TWB_NAME)
with open(twb_path, "w", encoding="utf-8") as f:
    f.write(twb_xml)
print(f"[OK] TWB guardado: {twb_path}")

# ─────────────────────────────────────────────────────────────────────────────
# 5.  EMPAQUETAR COMO .twbx
# ─────────────────────────────────────────────────────────────────────────────
twbx_path = os.path.join(OUT_DIR, TWBX_NAME)
with zipfile.ZipFile(twbx_path, "w", zipfile.ZIP_DEFLATED) as zf:
    # El workbook en raíz del zip (con el mismo nombre sin ext → Tableau lo encuentra)
    zf.write(twb_path, arcname=TWB_NAME)
    # El Excel en la carpeta Data/
    zf.write(excel_path, arcname=f"Data/{EXCEL_NAME}")

print(f"[OK] TWBX guardado: {twbx_path}")
print()
print("=" * 65)
print("  PROYECTO TABLEAU COMPLETADO")
print("  Abre  Mundiales_Confederaciones.twbx  en Tableau Desktop")
print("=" * 65)
print()

# Mostrar resumen de datos generados
print("RESUMEN DE DATOS:")
print(f"  Ediciones:        {len(df_ediciones)} mundiales (1930-2022)")
print(f"  Participaciones:  {len(df_part)} registros pais-edicion")
print(f"  Paises unicos:    {df_part['País'].nunique()}")
print(f"  Confederaciones:  {df_part['Confederación'].nunique()}")
print()
print("TITULOS POR CONFEDERACION:")
for _, r in df_rconf[df_rconf["Campeones"]>0].iterrows():
    print(f"  {r['Confederación']:10s}: {int(r['Campeones'])} titulos | "
          f"{int(r['N_Participaciones'])} participaciones")
print()
print("ARCHIVOS GENERADOS:")
print(f"  {EXCEL_NAME}")
print(f"  {TWB_NAME}")
print(f"  {TWBX_NAME}")
